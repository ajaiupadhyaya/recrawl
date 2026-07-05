"""
workbook_layout.py

Single source of truth for the U.S. Economy Dashboard workbook structure.
Both the one-time manual build and the weekly automated refresh
(refresh_economics_model.py) call build_workbook(data) with a data dict
in the same shape -- so the two never drift out of sync, and updating the
layout only ever has to happen in one place.

`data` schema (all keys required unless noted):
{
  "last_refreshed": str,
  "next_refresh": str,
  "snapshot": {                      # Macro Dashboard tab
      "gdp": (value, period, prior_note),
      "unemployment": (...), "payrolls": (...), "lfpr": (...), "jolts": (...),
      "cpi": (...), "corecpi": (...), "pce": (...), "corepce": (...),
      "fedfunds": (...), "y10": (...), "y2": (...),
  },
  "payroll_trend": {month_label: value_in_thousands, ...},   # 6 months, Labor Market tab
  "labor_internals": {stat_label: value, ...},               # Labor Market tab
  "inflation_trend": {                                        # Inflation tab, 3 months
      "cpi": {month: value}, "corecpi": {...}, "pce": {...}, "corepce": {...}
  },
  "yield_curve": {"1M": v, "3M": v, ..., "30Y": v},           # Inflation tab
  "gasoline": {"price_pre": v, "price_peak": v, "qty_pre": v, "qty_peak": v},
  "phillips_curve": {year: (unemployment, cpi_inflation), ...},  # 16 years
}
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, ScatterChart, Series, Reference

NAVY = "1F2937"
BLUE_TXT = "0000FF"
BLACK_TXT = "000000"
GREEN_TXT = "008000"
WHITE = "FFFFFF"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
REFRESH_FILL = PatternFill("solid", fgColor="E2EFDA")
INPUT_FONT = Font(name="Arial", size=10, color=BLUE_TXT)
CALC_FONT = Font(name="Arial", size=10, color=BLACK_TXT)
LINK_FONT = Font(name="Arial", size=10, color=GREEN_TXT)
BOLD_CALC = Font(name="Arial", size=10, color=BLACK_TXT, bold=True)
HDR_FONT = Font(name="Arial", size=12, color=WHITE, bold=True)
SUBHDR_FONT = Font(name="Arial", size=10, bold=True, color=BLACK_TXT)
TITLE_FONT = Font(name="Arial", size=18, bold=True, color=NAVY)
SOURCE_FONT = Font(name="Arial", size=8, italic=True, color="808080")

PCT1 = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
NUM0 = '#,##0;(#,##0);"-"'
MULT = '0.00'


def style_header(ws, row, col_start, col_end, text=None):
    if text:
        ws.cell(row=row, column=col_start, value=text)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_workbook(data):
    """Build and return the full 7-tab workbook from a data dict. See module
    docstring for the exact schema. Every blue-input cell below reads from
    `data` rather than a hardcoded literal, so refresh_economics_model.py
    can feed it live-pulled FRED/EIA values and get an identical layout."""
    wb = Workbook()

    # ---------------- COVER ----------------
    cov = wb.active
    cov.title = "Cover"
    set_widths(cov, {"A": 4, "B": 42, "C": 22, "D": 22, "E": 22, "F": 22})
    cov.sheet_view.showGridLines = False
    cov["B2"] = "U.S. ECONOMY DASHBOARD"
    cov["B2"].font = TITLE_FONT
    cov["B3"] = "Macro + Micro Analysis — auto-refreshed weekly"
    cov["B3"].font = Font(size=12, italic=True, color="444444")
    r = 6
    style_header(cov, r, 2, 6, "LAST DATA REFRESH"); r += 1
    cov.cell(row=r, column=2, value="Last Refreshed").font = CALC_FONT
    cov.cell(row=r, column=3, value=data["last_refreshed"]).font = INPUT_FONT
    cov.cell(row=r, column=3).fill = REFRESH_FILL
    r += 1
    cov.cell(row=r, column=2, value="Next Scheduled Refresh").font = CALC_FONT
    cov.cell(row=r, column=3, value=data["next_refresh"]).font = INPUT_FONT
    r += 2

    # ---------------- MACRO DASHBOARD ----------------
    md = wb.create_sheet("Macro Dashboard")
    set_widths(md, {"A": 2, "B": 40, "C": 16, "D": 16, "E": 30})
    md.sheet_view.showGridLines = False
    style_header(md, 2, 2, 5, "U.S. BUSINESS CYCLE SNAPSHOT")
    r = 3
    for h, col in zip(["Indicator", "Latest Value", "Reference Period", "Prior"], range(2, 6)):
        md.cell(row=r, column=col, value=h).font = SUBHDR_FONT
        md.cell(row=r, column=col).fill = SUBHDR_FILL
    r += 1
    snap = data["snapshot"]
    label_map = [
        ("gdp", "Real GDP Growth (annualized, QoQ)", PCT1),
        ("unemployment", "Unemployment Rate (U-3)", PCT1),
        ("payrolls", "Nonfarm Payrolls, Monthly Change", NUM0),
        ("lfpr", "Labor Force Participation Rate", PCT1),
        ("jolts", "JOLTS Job Openings", NUM0),
        ("cpi", "CPI Inflation, YoY (headline)", PCT1),
        ("corecpi", "Core CPI Inflation, YoY", PCT1),
        ("pce", "PCE Inflation, YoY", PCT1),
        ("corepce", "Core PCE Inflation, YoY", PCT1),
        ("fedfunds", "Fed Funds Rate (effective)", PCT2),
        ("y10", "10-Year Treasury Yield", PCT2),
        ("y2", "2-Year Treasury Yield", PCT2),
    ]
    md_rows = {}
    for key, label, fmt in label_map:
        val, period, prior = snap[key]
        md.cell(row=r, column=2, value=label).font = CALC_FONT
        md.cell(row=r, column=3, value=val).font = INPUT_FONT
        md.cell(row=r, column=3).number_format = fmt
        md.cell(row=r, column=3).fill = REFRESH_FILL
        md.cell(row=r, column=4, value=period).font = Font(size=9, italic=True)
        if prior:
            md.cell(row=r, column=5, value=prior).font = Font(size=9, italic=True)
        md_rows[key] = r
        r += 1
    r += 1
    md.cell(row=r, column=2, value="10Y-2Y Treasury Spread").font = Font(bold=True)
    md.cell(row=r, column=3, value=f"=C{md_rows['y10']}-C{md_rows['y2']}").font = BOLD_CALC
    md.cell(row=r, column=3).number_format = PCT2
    r += 1
    md.cell(row=r, column=2, value="Real Fed Funds Rate (Fed Funds - Core PCE)").font = Font(bold=True)
    md.cell(row=r, column=3, value=f"=C{md_rows['fedfunds']}-C{md_rows['corepce']}").font = BOLD_CALC
    md.cell(row=r, column=3).number_format = PCT2

    # ---------------- LABOR MARKET ----------------
    lm = wb.create_sheet("Labor Market")
    set_widths(lm, {"A": 2, "B": 30, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12})
    lm.sheet_view.showGridLines = False
    style_header(lm, 2, 2, 8, "NONFARM PAYROLLS -- MONTHLY CHANGE (thousands)")
    r = 3
    months = list(data["payroll_trend"].keys())
    for i, m in enumerate(months):
        lm.cell(row=r, column=3 + i, value=m).font = SUBHDR_FONT
        lm.cell(row=r, column=3 + i).fill = SUBHDR_FILL
    r += 1
    row_payroll = r
    for i, m in enumerate(months):
        lm.cell(row=r, column=3 + i, value=data["payroll_trend"][m]).font = INPUT_FONT
        lm.cell(row=r, column=3 + i).number_format = NUM0
        lm.cell(row=r, column=3 + i).fill = REFRESH_FILL
    lm.cell(row=r, column=2, value="Payrolls Added (thousands)").font = CALC_FONT
    r += 2
    chart = BarChart()
    chart.title = "Nonfarm Payrolls: Monthly Change (000s)"
    data_ref = Reference(lm, min_col=3, max_col=2 + len(months), min_row=row_payroll, max_row=row_payroll)
    cats = Reference(lm, min_col=3, max_col=2 + len(months), min_row=3, max_row=3)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats)
    chart.height, chart.width = 7, 14
    lm.add_chart(chart, f"B{r}")
    r += 16
    style_header(lm, r, 2, 8, "LABOR MARKET INTERNALS"); r += 1
    for label, val in data["labor_internals"].items():
        lm.cell(row=r, column=2, value=label).font = CALC_FONT
        lm.cell(row=r, column=3, value=val).font = INPUT_FONT
        lm.cell(row=r, column=3).fill = REFRESH_FILL
        r += 1

    # ---------------- INFLATION & MONETARY POLICY ----------------
    im = wb.create_sheet("Inflation & Monetary Policy")
    set_widths(im, {"A": 2, "B": 22, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13, "H": 13})
    im.sheet_view.showGridLines = False
    style_header(im, 2, 2, 6, "INFLATION TREND: HEADLINE vs. CORE")
    r = 3
    inf = data["inflation_trend"]
    inf_months = list(inf["cpi"].keys())
    for i, m in enumerate(inf_months):
        im.cell(row=r, column=3 + i, value=m).font = SUBHDR_FONT
        im.cell(row=r, column=3 + i).fill = SUBHDR_FILL
    r += 1
    for series_key, label in [("cpi", "CPI, YoY"), ("corecpi", "Core CPI, YoY"),
                               ("pce", "PCE, YoY"), ("corepce", "Core PCE, YoY")]:
        im.cell(row=r, column=2, value=label).font = CALC_FONT
        for i, m in enumerate(inf_months):
            im.cell(row=r, column=3 + i, value=inf[series_key][m]).font = INPUT_FONT
            im.cell(row=r, column=3 + i).number_format = PCT1
            im.cell(row=r, column=3 + i).fill = REFRESH_FILL
        r += 1
    r += 1
    style_header(im, r, 2, 6, "YIELD CURVE"); r += 1
    yc = data["yield_curve"]
    maturities = list(yc.keys())
    for i, m in enumerate(maturities):
        im.cell(row=r, column=3 + i, value=m).font = SUBHDR_FONT
        im.cell(row=r, column=3 + i).fill = SUBHDR_FILL
    r += 1
    for i, m in enumerate(maturities):
        im.cell(row=r, column=3 + i, value=yc[m]).font = INPUT_FONT
        im.cell(row=r, column=3 + i).number_format = PCT2
        im.cell(row=r, column=3 + i).fill = REFRESH_FILL

    # ---------------- MICRO: GASOLINE ELASTICITY ----------------
    ge = wb.create_sheet("Micro- Gasoline Elasticity")
    set_widths(ge, {"A": 2, "B": 42, "C": 16, "D": 16})
    ge.sheet_view.showGridLines = False
    style_header(ge, 2, 2, 4, "GASOLINE PRICE ELASTICITY OF DEMAND -- LIVE NATURAL EXPERIMENT")
    r = 4
    gas = data["gasoline"]
    ge.cell(row=r, column=2, value="U.S. Avg. Retail Gasoline Price ($/gal)").font = CALC_FONT
    ge.cell(row=r, column=3, value=gas["price_pre"]).font = INPUT_FONT
    ge.cell(row=r, column=3).number_format = CUR2
    ge.cell(row=r, column=4, value=gas["price_peak"]).font = INPUT_FONT
    ge.cell(row=r, column=4).number_format = CUR2
    row_p1 = r; r += 1
    ge.cell(row=r, column=2, value="Quantity Index (Qd proxy)").font = CALC_FONT
    ge.cell(row=r, column=3, value=gas["qty_pre"]).font = INPUT_FONT
    ge.cell(row=r, column=4, value=gas["qty_peak"]).font = INPUT_FONT
    row_q1 = r; r += 2
    row_pctq = r
    ge.cell(row=r, column=2, value="% Change in Quantity (arc method)").font = CALC_FONT
    ge.cell(row=r, column=3, value=f"=(D{row_q1}-C{row_q1})/((C{row_q1}+D{row_q1})/2)").font = CALC_FONT
    ge.cell(row=r, column=3).number_format = PCT1
    r += 1
    row_pctp = r
    ge.cell(row=r, column=2, value="% Change in Price (arc method)").font = CALC_FONT
    ge.cell(row=r, column=3, value=f"=(D{row_p1}-C{row_p1})/((C{row_p1}+D{row_p1})/2)").font = CALC_FONT
    ge.cell(row=r, column=3).number_format = PCT1
    r += 1
    ge.cell(row=r, column=2, value="PRICE ELASTICITY OF DEMAND").font = Font(bold=True, color=WHITE)
    ge.cell(row=r, column=2).fill = HDR_FILL
    ge.cell(row=r, column=3, value=f"=C{row_pctq}/C{row_pctp}").font = Font(bold=True, color=WHITE)
    ge.cell(row=r, column=3).fill = HDR_FILL
    ge.cell(row=r, column=3).number_format = '0.00'

    # ---------------- PHILLIPS CURVE REGRESSION ----------------
    pc = wb.create_sheet("Phillips Curve Regression")
    set_widths(pc, {"A": 2, "B": 10, "C": 14, "D": 14})
    pc.sheet_view.showGridLines = False
    style_header(pc, 2, 2, 4, "PHILLIPS CURVE -- LIVE REGRESSION")
    r = 4
    for h, col in zip(["Year", "Unemployment", "CPI Inflation"], range(2, 5)):
        pc.cell(row=r, column=col, value=h).font = SUBHDR_FONT
        pc.cell(row=r, column=col).fill = SUBHDR_FILL
    r += 1
    data_start = r
    for year, (unemp, cpi) in sorted(data["phillips_curve"].items()):
        pc.cell(row=r, column=2, value=year).font = INPUT_FONT
        pc.cell(row=r, column=3, value=unemp).font = INPUT_FONT
        pc.cell(row=r, column=3).number_format = PCT1
        pc.cell(row=r, column=4, value=cpi).font = INPUT_FONT
        pc.cell(row=r, column=4).number_format = PCT1
        r += 1
    data_end = r - 1
    r += 1
    pc.cell(row=r, column=2, value="Slope").font = Font(bold=True)
    pc.cell(row=r, column=3, value=f"=SLOPE(D{data_start}:D{data_end},C{data_start}:C{data_end})").font = BOLD_CALC
    r += 1
    pc.cell(row=r, column=2, value="R-Squared").font = Font(bold=True)
    pc.cell(row=r, column=3, value=f"=RSQ(D{data_start}:D{data_end},C{data_start}:C{data_end})").font = BOLD_CALC
    r += 2
    chart = ScatterChart()
    chart.title = "Unemployment vs Inflation"
    xvalues = Reference(pc, min_col=3, min_row=data_start, max_row=data_end)
    yvalues = Reference(pc, min_col=4, min_row=data_start, max_row=data_end)
    series = Series(yvalues, xvalues, title="Unemp vs Inflation")
    series.marker.symbol = "circle"
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.height, chart.width = 9, 15
    pc.add_chart(chart, f"F{data_start}")

    # ---------------- AUTOMATION & SOURCES ----------------
    au = wb.create_sheet("Automation & Sources")
    set_widths(au, {"A": 2, "B": 60})
    au.sheet_view.showGridLines = False
    style_header(au, 2, 2, 3, "AUTOMATION LOG")
    au.cell(row=4, column=2, value=f"Last refreshed: {data['last_refreshed']}").font = CALC_FONT
    au.cell(row=5, column=2, value="See project README for the full FRED/EIA series list and setup steps.").font = SOURCE_FONT

    wb.active = 0
    return wb
